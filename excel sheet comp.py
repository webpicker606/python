import openpyxl as xl

wb = xl.load_workbook('excel.xlsx')
sheet = wb['Sheet1']
cell = sheet.cell(1, 1)
for row in range(2, sheet.max_row+1):
    cell = sheet.cell(row, 3)
    corrected_value = cell.value*0.90
    corrected_cell = sheet.cell(row, 4)
    corrected_cell.value = corrected_value

wb.save('book2.xlsx')
